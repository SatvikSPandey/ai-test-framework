import uuid
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from jinja2 import Environment, FileSystemLoader
from core.models import (
    ParsedRequirements, GeneratedTestCases, GeneratedScripts,
    ExecutionResults, FinalReport, CoverageItem, TestStatus
)
from core.config import settings


class ReportGenerator:
    """
    Agent 5 — Report Generator.
    Takes outputs from all previous agents.
    Produces three report formats:
    1. HTML dashboard — visual, with pass/fail charts and screenshots
    2. Excel file — five sheets, one per agent output
    3. JSON file — machine readable, for programmatic consumption
    Calculates requirement coverage — which features have passing tests.
    """

    def __init__(self):
        self.run_id = str(uuid.uuid4())[:8].upper()
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.report_dir = settings.outputs_dir / f"run_{self.timestamp}"
        self.report_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        parsed_requirements: ParsedRequirements,
        generated_test_cases: GeneratedTestCases,
        generated_scripts: GeneratedScripts,
        execution_results: ExecutionResults
    ) -> FinalReport:
        """
        Main entry point. Generates all three report formats.
        Returns a FinalReport object with paths to all generated files.
        """
        print("  Building coverage map...")
        coverage_map = self._build_coverage_map(
            parsed_requirements, generated_test_cases, execution_results
        )

        coverage_percentage = (
            sum(1 for c in coverage_map if c.covered) / len(coverage_map) * 100
            if coverage_map else 0.0
        )

        print("  Generating Excel report...")
        excel_path = self._generate_excel(
            parsed_requirements, generated_test_cases,
            generated_scripts, execution_results, coverage_map
        )

        print("  Generating JSON report...")
        json_path = self._generate_json(
            parsed_requirements, generated_test_cases,
            execution_results, coverage_map, coverage_percentage
        )

        print("  Generating HTML report...")
        html_path = self._generate_html(
            parsed_requirements, generated_test_cases,
            execution_results, coverage_map, coverage_percentage
        )

        return FinalReport(
            run_id=self.run_id,
            target_url=settings.target_url,
            total_features=parsed_requirements.total_features,
            total_test_cases=generated_test_cases.total_test_cases,
            total_passed=execution_results.passed,
            total_failed=execution_results.failed,
            coverage_percentage=round(coverage_percentage, 1),
            coverage_map=coverage_map,
            html_report_path=str(html_path),
            excel_report_path=str(excel_path),
            json_report_path=str(json_path)
        )

    def _build_coverage_map(
        self,
        parsed_requirements: ParsedRequirements,
        generated_test_cases: GeneratedTestCases,
        execution_results: ExecutionResults
    ) -> list[CoverageItem]:
        """
        For each feature, counts how many test cases were written
        and how many passed. Marks a feature as covered if at least
        one test case passed.
        """
        # Build a lookup of test results by test case id
        results_by_id = {
            r.test_case_id: r for r in execution_results.results
        }

        coverage_map = []

        for feature in parsed_requirements.features:
            # Find all test cases for this feature
            feature_tests = [
                tc for tc in generated_test_cases.test_cases
                if tc.feature_id == feature.id
            ]

            # Count passing tests for this feature
            passing = sum(
                1 for tc in feature_tests
                if tc.id in results_by_id
                and results_by_id[tc.id].status == TestStatus.PASSED
            )

            coverage_map.append(CoverageItem(
                feature_id=feature.id,
                feature_name=feature.name,
                total_criteria=len(feature.acceptance_criteria),
                tests_written=len(feature_tests),
                tests_passed=passing,
                covered=passing > 0
            ))

        return coverage_map

    def _generate_excel(
        self,
        parsed_requirements, generated_test_cases,
        generated_scripts, execution_results, coverage_map
    ) -> Path:
        """Generates an Excel file with five sheets — one per agent output."""
        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E4057")

        ws.append(["AI Test Framework — Run Summary"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Run ID", self.run_id])
        ws.append(["Timestamp", self.timestamp])
        ws.append(["Target URL", settings.target_url])
        ws.append(["Source File", parsed_requirements.source_file])
        ws.append([])
        ws.append(["Total Features", parsed_requirements.total_features])
        ws.append(["Total Test Cases", generated_test_cases.total_test_cases])
        ws.append(["Passed", execution_results.passed])
        ws.append(["Failed", execution_results.failed])
        ws.append(["Errors", execution_results.errors])
        ws.append(["Skipped", execution_results.skipped])
        ws.append([])
        covered = sum(1 for c in coverage_map if c.covered)
        ws.append(["Features Covered", f"{covered}/{len(coverage_map)}"])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        # ── Sheet 2: Requirements ─────────────────────────────────────────
        ws2 = wb.create_sheet("Requirements")
        headers = ["Feature ID", "Feature Name", "Description", "Acceptance Criteria Count"]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for feature in parsed_requirements.features:
            ws2.append([
                feature.id,
                feature.name,
                feature.description,
                len(feature.acceptance_criteria)
            ])
        for col in ["A", "B", "C", "D"]:
            ws2.column_dimensions[col].width = 30

        # ── Sheet 3: Test Cases ───────────────────────────────────────────
        ws3 = wb.create_sheet("Test Cases")
        headers = ["Test ID", "Feature ID", "Title", "Priority", "Steps", "Confidence"]
        ws3.append(headers)
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
        for tc in generated_test_cases.test_cases:
            ws3.append([
                tc.id,
                tc.feature_id,
                tc.title,
                tc.priority.value,
                len(tc.steps),
                tc.confidence_score
            ])
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws3.column_dimensions[col].width = 25

        # ── Sheet 4: Execution Results ────────────────────────────────────
        ws4 = wb.create_sheet("Execution Results")
        headers = ["Test ID", "Status", "Duration (s)", "Attempts", "Error"]
        ws4.append(headers)
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
        for result in execution_results.results:
            ws4.append([
                result.test_case_id,
                result.status.value,
                result.duration_seconds,
                result.attempts,
                result.error_message or ""
            ])
        for col in ["A", "B", "C", "D", "E"]:
            ws4.column_dimensions[col].width = 25

        # ── Sheet 5: Coverage Map ─────────────────────────────────────────
        ws5 = wb.create_sheet("Coverage Map")
        headers = ["Feature ID", "Feature Name", "Criteria", "Tests Written", "Tests Passed", "Covered"]
        ws5.append(headers)
        for cell in ws5[1]:
            cell.font = header_font
            cell.fill = header_fill
        for item in coverage_map:
            ws5.append([
                item.feature_id,
                item.feature_name,
                item.total_criteria,
                item.tests_written,
                item.tests_passed,
                "YES" if item.covered else "NO"
            ])
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws5.column_dimensions[col].width = 25

        excel_path = self.report_dir / f"report_{self.timestamp}.xlsx"
        wb.save(str(excel_path))
        return excel_path

    def _generate_json(
        self, parsed_requirements, generated_test_cases,
        execution_results, coverage_map, coverage_percentage
    ) -> Path:
        """Generates a structured JSON report for programmatic consumption."""
        report = {
            "run_id": self.run_id,
            "timestamp": self.timestamp,
            "target_url": settings.target_url,
            "source_file": parsed_requirements.source_file,
            "summary": {
                "total_features": parsed_requirements.total_features,
                "total_test_cases": generated_test_cases.total_test_cases,
                "passed": execution_results.passed,
                "failed": execution_results.failed,
                "errors": execution_results.errors,
                "skipped": execution_results.skipped,
                "coverage_percentage": round(coverage_percentage, 1)
            },
            "coverage_map": [
                {
                    "feature_id": c.feature_id,
                    "feature_name": c.feature_name,
                    "tests_written": c.tests_written,
                    "tests_passed": c.tests_passed,
                    "covered": c.covered
                }
                for c in coverage_map
            ],
            "results": [
                {
                    "test_case_id": r.test_case_id,
                    "status": r.status.value,
                    "duration_seconds": r.duration_seconds,
                    "attempts": r.attempts,
                    "error_message": r.error_message
                }
                for r in execution_results.results
            ]
        }

        json_path = self.report_dir / f"report_{self.timestamp}.json"
        json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
        return json_path

    def _generate_html(
        self, parsed_requirements, generated_test_cases,
        execution_results, coverage_map, coverage_percentage
    ) -> Path:
        """
        Generates an HTML dashboard report.
        Uses a Jinja2 template from the templates/ folder.
        Falls back to a basic HTML string if the template is missing.
        """
        context = {
            "run_id": self.run_id,
            "timestamp": self.timestamp,
            "target_url": settings.target_url,
            "source_file": parsed_requirements.source_file,
            "total_features": parsed_requirements.total_features,
            "total_test_cases": generated_test_cases.total_test_cases,
            "passed": execution_results.passed,
            "failed": execution_results.failed,
            "errors": execution_results.errors,
            "skipped": execution_results.skipped,
            "coverage_percentage": round(coverage_percentage, 1),
            "coverage_map": coverage_map,
            "results": execution_results.results
        }

        template_path = settings.templates_dir / "report.html"

        if template_path.exists():
            env = Environment(loader=FileSystemLoader(str(settings.templates_dir)))
            template = env.get_template("report.html")
            html_content = template.render(**context)
        else:
            # Basic fallback HTML if template file doesn't exist yet
            html_content = self._basic_html(context)

        html_path = self.report_dir / f"report_{self.timestamp}.html"
        html_path.write_text(html_content, encoding="utf-8")
        return html_path

    def _basic_html(self, ctx: dict) -> str:
        """Basic HTML fallback when Jinja2 template is not available."""
        return f"""<!DOCTYPE html>
<html>
<head><title>AI Test Framework Report</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
  h1 {{ color: #2E4057; }}
  .summary {{ background: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
  .passed {{ color: green; font-weight: bold; }}
  .failed {{ color: red; font-weight: bold; }}
  table {{ width: 100%; border-collapse: collapse; background: white; }}
  th {{ background: #2E4057; color: white; padding: 10px; text-align: left; }}
  td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
</style>
</head>
<body>
<h1>AI Test Framework — Run Report</h1>
<div class="summary">
  <p><strong>Run ID:</strong> {ctx['run_id']}</p>
  <p><strong>Timestamp:</strong> {ctx['timestamp']}</p>
  <p><strong>Target URL:</strong> {ctx['target_url']}</p>
  <p><strong>Coverage:</strong> {ctx['coverage_percentage']}%</p>
  <p><span class="passed">Passed: {ctx['passed']}</span> &nbsp;
     <span class="failed">Failed: {ctx['failed']}</span> &nbsp;
     Errors: {ctx['errors']} &nbsp; Skipped: {ctx['skipped']}</p>
</div>
<h2>Execution Results</h2>
<table>
  <tr><th>Test ID</th><th>Status</th><th>Duration</th><th>Attempts</th><th>Error</th></tr>
  {''.join(f"<tr><td>{r.test_case_id}</td><td>{r.status.value}</td><td>{r.duration_seconds}s</td><td>{r.attempts}</td><td>{r.error_message or ''}</td></tr>" for r in ctx['results'])}
</table>
</body>
</html>"""


# Single shared instance
report_generator = ReportGenerator()